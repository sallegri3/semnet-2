import pandas as pd
from openpyxl import load_workbook
import os

#
# def write_multisheet_excel(sheets, sheet_names, filepath, overwrite=True):
#     # Write results to spreadsheet
#     # path = 'gene_orch_rankings.xlsx'
#     # if os.path.isfile(path):
#     #     book = load_workbook(path)
#     #     writer = pd.ExcelWriter(path, engine = 'openpyxl')
#     #     writer.book = book
#     # else:
#     #     writer = pd.ExcelWriter(path, engine = 'openpyxl')
#     with pd.ExcelWriter(filepath) as writer:
#         for sheet, name in zip(sheets, sheet_names):
#             print(name)
#             sheet.to_excel(writer, sheet_name=name.title())
#             writer.save()
#     writer.close()
#
def write_multisheet_excel(sheets, sheet_names, filepath, overwrite=True):
    '''
    Write multiple dataframes to a single excel workbook as multiple sheets

    Inputs:
        sheets: list of pandas.DataFrame
            List of dataframes you want to put in an excel notebook

        sheet_names: list of string
            Names of sheets for excel notebook

        filepath: string
            Path to write excel notebook

        overwrite: bool
            Whether to overwrite a previous file with same name
    '''
    if overwrite:
        if os.path.isfile(filepath):
            os.remove(filepath)
    # Write results to spreadsheet
    for sheet, name in zip(sheets, sheet_names):
        if os.path.isfile(filepath):
            book = load_workbook(filepath)
            writer = pd.ExcelWriter(filepath, engine = 'openpyxl')
            writer.book = book
        else:
            writer = pd.ExcelWriter(filepath, engine = 'openpyxl')
        sheet.to_excel(writer, sheet_name=name)
        writer.save()
